from sqlalchemy.orm import Session
from typing import List, Optional, Dict, Any
import asyncio
import logging
import pandas as pd
import tempfile
import os
from datetime import datetime
from openpyxl.styles import PatternFill
from db.models import GTSample, CourseCatalog
from schemas.gt_sample import GTSampleCreate
from util.ingestion_client import IngestionClient
from util.profile_converter import _create_course_info_dict

logger = logging.getLogger(__name__)


def _create_profile_query(
    query: str,
    major: str, 
    interest_job: str,  
    certification: str,
    club_activities: str, 
    course_info_dicts: List[Dict[str, Any]] = []
) -> str:
    """
    GT 프로필 정보로부터 완전한 검색 쿼리 생성
    (ingestion/src/services/query_builder.py의 ProfileQueryBuilder.create_profile_query와 동일한 로직)
    """
    query_parts = []

    # 1. 사용자 질문
    if query:
        query_parts.append(f"질문: {query}")

    # 2. 전공
    if major:
        query_parts.append(f"전공: {major}")

    # 3. 관심 직무 (문자열을 리스트로 변환)
    if interest_job:
        interest_list = [item.strip() for item in interest_job.split(",") if item.strip()]
        if interest_list:
            query_parts.append(f"관심 직무: {', '.join(interest_list)}")

    # 4. 자격증
    if certification:
        query_parts.append(f"자격증: {certification}")

    # 5. 동아리/대외활동
    if club_activities:
        query_parts.append(f"동아리/대외활동: {club_activities}")

    # 6. 수강 이력
    if course_info_dicts:
        query_parts.append("수강 이력:")
        for course in course_info_dicts:
            course_parts = []
            course_parts.append(f"강의명: {course.get('course_name', '')}")
            course_parts.append(f"핵심 역량: {course.get('core_competency', '')}")
            course_parts.append(f"강의 개요: {course.get('course_overview', '')}")
            course_parts.append(f"학습 목표: {course.get('course_objectives', '')}")
            
            # 주차별 계획 (1-16주차)
            week_plans = []
            for i in range(1, 17):
                week_plan = course.get(f'week{i}_plan', '')
                week_plans.append(f"{i}주차: {week_plan}")
            
            course_parts.append(f"주차별 계획: {' / '.join(week_plans)}")
            
            # 강의 정보들을 " | "로 구분하여 하나의 라인으로
            query_parts.append(" | ".join(course_parts))
    
    # 모든 부분을 줄바꿈으로 구분
    return "\n".join(query_parts)


def create_gt_sample(db: Session, data: GTSampleCreate) -> GTSample:
    # RelevantDoc 객체를 JSON 직렬화 가능한 dict 로 변환
    rel = [item.dict() if hasattr(item, "dict") else item for item in data.relevant_ids]

    sample = GTSample(
        seed_rec_idx=data.seed_rec_idx,
        relevant_ids=rel,
        profile=data.profile,
        query=data.query,
    )
    db.add(sample)
    db.commit()
    db.refresh(sample)
    return sample


def get_gt_sample(db: Session, sample_id: int) -> Optional[GTSample]:
    return db.query(GTSample).filter(GTSample.id == sample_id).first()


def list_gt_samples(db: Session, skip: int = 0, limit: int = 200) -> List[GTSample]:
    return db.query(GTSample).offset(skip).limit(limit).all()


async def enrich_all_gt_metadata(db: Session) -> Dict[str, int]:
    """모든 GT 샘플의 메타데이터 수집 및 업데이트"""
    
    all_samples = db.query(GTSample).all()
    logger.info(f"총 {len(all_samples)}개 GT 샘플 발견")
    
    ingestion_client = IngestionClient()
    success_count = 0
    failed_count = 0
    
    for sample in all_samples:
        try:
            relevant_ids = sample.relevant_ids
            docs_metadata = []
            
            for rel_doc in relevant_ids:
                rec_idx = rel_doc.get("rec_idx")
                distance = rel_doc.get("distance", 0.0)
                
                if not rec_idx:
                    continue
                
                posting_data = await ingestion_client.get_posting(rec_idx)
                
                if posting_data:
                    metadata = posting_data.get("metadata", {})
                    excerpt = posting_data.get("excerpt", "")
                    
                    doc_meta = {
                        "rec_idx": rec_idx,
                        "distance": distance,
                        "title": metadata.get("post_title", metadata.get("title", "제목 없음")),
                        "company": metadata.get("company", "회사명 없음"),
                        "deadline": metadata.get("deadline", "마감일 미정"),
                        "url": metadata.get("detail_url", ""),
                        "excerpt": excerpt[:200] if excerpt else "내용 없음"
                    }
                else:
                    doc_meta = {
                        "rec_idx": rec_idx,
                        "distance": distance,
                        "title": "조회 실패",
                        "company": "Unknown",
                        "deadline": "Unknown",
                        "url": "",
                        "excerpt": "메타데이터 조회에 실패했습니다."
                    }
                
                docs_metadata.append(doc_meta)
            
            sample.relevant_docs_metadata = {"docs": docs_metadata}
            success_count += 1
            logger.info(f"GT 샘플 {sample.id} 메타데이터 보강 완료")
            
            await asyncio.sleep(0.3)
            
        except Exception as e:
            failed_count += 1
            logger.error(f"GT 샘플 {sample.id} 처리 실패: {e}")
    
    try:
        db.commit()
        logger.info("모든 메타데이터 업데이트 커밋 완료")
    except Exception as e:
        db.rollback()
        logger.error(f"커밋 실패: {e}")
        raise
    
    return {
        "total": len(all_samples),
        "success": success_count,
        "failed": failed_count
    }


def export_gt_data_to_excel(db: Session, version: Optional[str] = None) -> str:
    """GT 데이터를 Excel 파일로 export하고 임시 파일 경로 반환"""
    
    # GT 데이터 조회
    gt_samples = db.query(GTSample).all()
    
    if not gt_samples:
        raise ValueError("GT 샘플이 없습니다")
    
    # 데이터 변환
    excel_data = []
    
    for sample in gt_samples:
        profile = sample.profile
        metadata = sample.relevant_docs_metadata or {"docs": []}
        
        # 수강과목 목록 생성 (course_name만 추출)
        catalogs = profile.get("catalogs", [])
        course_names = [catalog.get("course_name", "") for catalog in catalogs if catalog.get("course_name")]
        course_list = ", ".join(course_names)
        
        # 완전한 검색 쿼리 생성
        complete_search_query = ""
        try:
            # Course ID들로 상세 정보 조회
            course_ids = [catalog.get("id") for catalog in catalogs if catalog.get("id")]
            if course_ids:
                course_details = db.query(CourseCatalog).filter(
                    CourseCatalog.id.in_(course_ids)
                ).all()
                
                # CourseInfo 딕셔너리로 변환
                course_info_dicts = [_create_course_info_dict(course) for course in course_details]
                
                # 완전한 검색 쿼리 생성
                complete_search_query = _create_profile_query(
                    query=sample.query,
                    major=profile.get("major", ""),
                    interest_job=profile.get("interest_job", ""),
                    certification=profile.get("certification", ""),
                    club_activities=profile.get("club_activities", ""),
                    course_info_dicts=course_info_dicts
                )
            else:
                complete_search_query = "수강과목 정보 없음"
                
        except Exception as e:
            logger.error(f"Sample {sample.id} 완전한 검색 쿼리 생성 실패: {e}")
            complete_search_query = "쿼리 생성 실패"
        
        # 관심분야 목록 생성 (타입 체크 추가)
        interest_jobs = profile.get("interest_job", [])
        if isinstance(interest_jobs, list):
            interest_list = ", ".join(interest_jobs)
        elif isinstance(interest_jobs, str):
            interest_list = interest_jobs
        else:
            interest_list = str(interest_jobs) if interest_jobs else ""
        
        # 관련 공고가 없는 경우 기본 행 추가
        docs = metadata.get("docs", [])
        if not docs:
            excel_data.append({
                "GT_ID": sample.id,
                "학생_전공": profile.get("major", ""),
                "학생_관심분야": interest_list,
                "수강과목": course_list,
                "완전한_검색_쿼리": complete_search_query,
                "학생_질문": sample.query,
                "공고_제목": "관련 공고 없음",
                "회사명": "",
                "거리점수": "",
                "URL": ""
            })
        else:
            # 각 관련 공고별로 행 생성
            for idx, doc in enumerate(docs):
                distance = doc.get("distance", 0)
                
                excel_data.append({
                    "GT_ID": sample.id,
                    "학생_전공": profile.get("major", ""),
                    "학생_관심분야": interest_list,
                    "수강과목": course_list,
                    "완전한_검색_쿼리": complete_search_query,
                    "학생_질문": sample.query,
                    "공고_제목": doc.get("title", ""),
                    "회사명": doc.get("company", ""),
                    "거리점수": round(distance, 3),
                    "URL": doc.get("url", "")
                })
    
    # DataFrame 생성
    df = pd.DataFrame(excel_data)
    
    # 임시 파일 생성
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    
    # Excel 파일 생성 및 스타일링
    with pd.ExcelWriter(temp_file.name, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='GT분석', index=False)
        
        workbook = writer.book
        worksheet = writer.sheets['GT분석']
        
        # 색상 정의
        color1 = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")  # 연한 파랑
        color2 = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")  # 연한 회색
        
        # GT_ID별로 색상 번갈아가며 적용
        current_gt_id = None
        color_index = 0
        
        for row_num, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
            gt_id = row[0].value  # GT_ID 컬럼
            
            # 새로운 GT_ID가 나타나면 색상 변경
            if gt_id != current_gt_id:
                current_gt_id = gt_id
                color_index = (color_index + 1) % 2
            
            # 현재 색상 선택
            fill_color = color1 if color_index == 0 else color2
            
            # 행 전체에 색상 적용
            for cell in row:
                cell.fill = fill_color
        
        # 컬럼 너비 자동 조정
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            column_name = column[0].value if column[0].value else ""
            if column_name in ['수강과목', '완전한_검색_쿼리', '학생_질문']:
                adjusted_width = min(max_length + 2, 80)
            else:
                adjusted_width = min(max_length + 2, 30)
            
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    logger.info(f"Excel 파일 생성 완료: {temp_file.name}")
    return temp_file.name 